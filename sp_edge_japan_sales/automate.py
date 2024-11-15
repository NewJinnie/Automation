import re
from playwright.sync_api import Playwright, sync_playwright, expect
import pyotp
import openpyxl
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from dotenv import load_dotenv
import sys
from common.netsuite_import import run as netsuite_import
from common.output_csv import run as output_csv
import pandas as pd
import xlwings as xw

# パスワードなどの環境変数のインポート
load_dotenv(".env")
OKTA_TOTP_SECRET_KEY = os.getenv("OKTA_TOTP_SECRET_KEY")
OKTA_PASSWORD = os.getenv("OKTA_PASSWORD")
okta_totp = pyotp.TOTP(OKTA_TOTP_SECRET_KEY)

# 該当年月の取得
dt_now = datetime.datetime.now()
year_str = str(dt_now.year)
month_str = str(dt_now.month - 1).zfill(2)
year_month_str = year_str + month_str

# 作業ファイル定義
work_file_folder = "C:\\Users\\shinya.arai\\Box\\経理\\12_Uzabase USA,Inc\\002_売上\\5_計上、前受振替\\" + year_str + "\\" + year_str + "." + month_str + "\\日本販売\\"
work_file_name = "SPEdge_日本販売_計上資料_" + year_month_str + ".xlsx"
work_file_path = work_file_folder + work_file_name
csv_sheet_num = 1

# 仕訳CSVの場所の定義
download_folder = "C:\\Users\\shinya.arai\\Downloads\\"
charge_je_csv_path = download_folder + "Edge請求仕訳" + year_month_str + ".csv"

# 仕訳CSVをNetsuiteにインポートする関数
def netsuite():
    with sync_playwright() as playwright:
        netsuite_import(playwright=playwright, file_path=charge_je_csv_path)

function_map = {
    "ns": netsuite,
}

# メイン関数
def run(playwright: Playwright) -> None:

    # ブラウザを立ち上げ、Salesforceへログイン
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://ub-saas.my.salesforce.com/")
    page.get_by_role("button", name="次を使用してログイン Okta").click()
    page.get_by_label("ユーザー名").click()
    page.get_by_label("ユーザー名").fill("shinya.arai@uzabase.com")
    page.get_by_role("button", name="次へ").click()
    page.get_by_label("パスワード").click()
    page.get_by_label("パスワード").fill(OKTA_PASSWORD)
    page.get_by_role("button", name="確認").click()
    page.get_by_label("コードを入力する").click()
    page.get_by_label("コードを入力する").fill(okta_totp.now())
    page.get_by_role("button", name="確認").click()

    # ログイン完了
    # 請求レポートをダウンロードする
    page.get_by_role("button", name="ナビゲーション項目をさらに表示").click()
    page.get_by_role("menuitem", name="レポート").click()
    page.get_by_role("link", name="[SP Edge JP]前受収益精査用_請求レポート_2024年9月まで").click()

    iframe_name = page.get_attribute("iframe[title=\"レポートビューアー\"]", "name")

    page.wait_for_selector(f"iframe[name=\"{iframe_name}\"]").content_frame().get_by_role("button", name="追加アクション").click()
    page.wait_for_selector(f"iframe[name=\"{iframe_name}\"]").content_frame().get_by_role("menuitem", name="エクスポート").click()
    page.get_by_text("詳細行のみをエクスポートします。これを使用して、さらなる計算や他のシステムへのアップロードを行います。").click()
    page.get_by_label("形式").select_option("xlsx")
    with page.expect_download() as download_info:
        with page.expect_popup() as page1_info:
            page.get_by_role("button", name="エクスポート").click()
        page1 = page1_info.value
    download = download_info.value
    download_path = download_folder + download.suggested_filename
    download.save_as(download_path)
    page1.close()

    # ダウンロード完了

    # pandasで読み込んで顧客コードでピボットする
    pd.set_option('display.unicode.east_asian_width', True)
    downloaded_df = pd.read_excel(download_path, sheet_name=0, index_col=None)
    headers = downloaded_df.columns
    indexes = [headers[1], headers[3]]
    values = headers[12]
    downloaded_df[values] = downloaded_df[values].astype(int) # 金額をInt型にする
    df_pivot = downloaded_df.pivot_table(index=indexes, values=values)
    
    # データを作業ファイルに貼付
    wb = openpyxl.load_workbook(work_file_path)
    ws_for_paste = wb[wb.sheetnames[0]]
    ws_for_paste.delete_rows(1, ws_for_paste.max_row)

    for row in dataframe_to_rows(df_pivot, index=True, header=True):
        ws_for_paste.append(row)

    # 上書き保存
    wb.save(work_file_path)

    # excel内の数式を動かすためにバックグラウンドでexcelを開いて保存して閉じる
    with xw.App(visible=False) as app:
        wb = app.books.open(work_file_path)
        wb.save()
        wb.close()

    # 仕訳CSVを出力する
    output_csv(work_file_path=work_file_path, sheet_num=csv_sheet_num, output_csv_path=charge_je_csv_path)

    # ---------------------
    context.close()
    browser.close()

if __name__ == "__main__":
    if len(sys.argv) > 1:
        function_name = sys.argv[1]
        func = function_map.get(function_name)
        if func:
            func()
    else:
        with sync_playwright() as playwright:
            run(playwright)